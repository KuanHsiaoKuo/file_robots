import openpyxl
import re
import xlrd
from datetime import datetime


def write_basic_excel(export_datas, template_path, output):
    valid_rows = {}
    for row_data in export_datas:
        if row_data[0].startswith('5'):
            num, title, _, _, current_debit, ending_debit, _, _ = row_data
            valid_rows[num] = {
                "title": title,
                # 本期贷方数
                'current_debit': float(str(current_debit).replace(',', '')) if current_debit else 0,
                # 本期借方数
                "ending_debit": float(str(ending_debit).replace(',', '')) if ending_debit else 0
            }
    template_wb = openpyxl.load_workbook(template_path)
    template_ws = template_wb['Sheet1']
    target_key = "ending_debit"
    for row_index, row_data in enumerate(template_ws):
        for col_index, cell in enumerate(row_data):
            if isinstance(cell.value, str) or isinstance(cell.value, int):
                cell_value = str(cell.value)
                try:
                    if "+" in cell_value and cell_value.startswith('5'):
                        total_list = [valid_rows[seg][target_key] for seg in cell_value.split("+")]
                        total = sum(total_list)
                        # 转为万元
                        cell.value = float("%.2f" % (total / 10000)) if total else 0
                    elif cell_value.startswith('5'):
                        cell.value = float("%.2f" % (valid_rows[cell_value][target_key] / 10000)) if valid_rows[cell_value][target_key] else 0
                    elif cell_value == "xxx":
                        cell.value = 0
                except Exception as e:
                    print(e.args[0])
                    print(cell.value)
            elif cell.value:
                print(cell.value, type(cell.value))
    template_wb.save(output)


def parse_project_data(export_datas):
    """
    1. 02[行政管理类]排除
    2. 分类为空的也不要
    3. 剩余分类按照项目+科目分类统计，有就填，无则为零。
    4. 结果按照项目分类、财政、事业收入这三个标准排序
    :param export_datas:
    :return:
    """
    exclude_category = ["02[行政管理类]"]
    headers = export_datas[1]
    valid_datas = []
    for row in export_datas[2:]:
        data = {headers[index]: cell for index, cell in enumerate(row)}
        for amount_key in ('借方金额', '贷方金额'):
            if not data[amount_key]:
                data[amount_key] = 0
            else:
                data[amount_key] = float(data[amount_key].replace(',', ''))
        invalid_conditions = [
            data['项目分类'] in exclude_category,
            not data['项目']
        ]
        if any(invalid_conditions):
            continue
        else:
            valid_datas.append(data)
    statistics_map = [
        ('项目', '项目'),
        ('项目分类', '项目分类'),
        ('财政项目收入', '')
    ]
    statistics_data = {}
    total_clns = []
    balance_key = '项目余额'
    expense_sum = '支出小计'
    profit_keys = ['财政项目收入', '事业收入']
    projects_result = {}
    for row in valid_datas:
        project = row['项目']
        class_names = re.findall(r"\[(.+)\]", row['科目'])
        cln = class_names[0] if class_names else row['摘要']
        if cln not in total_clns:
            total_clns.append(cln)
        # 1. 保存部门、项目支出小计，用于验算
        if cln == '部门、项目小计':
            if project not in projects_result.keys():
                projects_result[project] = {}
            projects_result[project]['借方金额'] = row['借方金额']
            projects_result[project]['余额'] = row['余额']
        else:
            debt_amount = row['借方金额']
            credit_amount = row['贷方金额']
            if project not in statistics_data.keys():
                statistics_data[project] = {balance_key: 0, expense_sum: 0}
            if cln not in statistics_data[project].keys():
                statistics_data[project][cln] = 0
            # 2. 统计科目
            statistics_data[project][cln] += debt_amount + credit_amount
            # 3. 统计项目余额
            statistics_data[project][balance_key] += debt_amount + credit_amount
            # 4. 统计支出小计
            if cln not in profit_keys:
                statistics_data[project][expense_sum] += debt_amount + credit_amount
            # 统一灌入剩余字段内容
            # for cell_key, cell_value in row.items():
            #     if cell_key not in statistics_data[project].keys():
            #         statistics_data[project][cell_key] = cell_value
    print(total_clns)
    result_headers = ['项目', '项目分类', balance_key]
    # 支出小计里的数，是十项费用求和, 等于明细里面部门、项目小计的借方金额debt_amount
    # 项目余额: 等于财政收入或事业收入减去支出小计, 核对为部门、项目小计的余额
    template_headers = ['项目', '项目分类', '财政项目收入', '事业收入', '办公费', '印刷资料费', '咨询费', '邮电费', '差旅费', '劳务费', '委托业务费', '税金', '间接费用', '其他商品和服务支出', '支出小计', balance_key]
    result_headers.extend(total_clns)
    result = [template_headers]
    for project, statistic in statistics_data.items():
        result_debt_amount, result_balance = projects_result[project]['借方金额'], projects_result[project]['余额']
        details = [statistic.get(cln, 0) for cln in template_headers]
        details.extend([result_debt_amount, result_balance])
        result.append(details)
    for line in result:
        print(line)
    return result, projects_result


def write_project_excel(export_datas, template_path, output_path):
    sum_data, projects_result = parse_project_data(export_datas)
    template_wb = openpyxl.load_workbook(template_path)
    template_ws = template_wb['项目']
    # template_wb = openpyxl.Workbook()
    # template_ws = template_wb.active
    # template_ws.title = "项目收支统计表"
    for row in sum_data:
        template_ws.append(row)
    template_wb.save(filename=output_path)
    print(f"{template_ws.title}保存在{output_path}")


operator_matches = {
    '总账余额表': write_basic_excel,
    '项目收支明细账': write_project_excel
}


def read_export_excel(file_path: str, template_path: str):
    path, extension = file_path.split('.')
    output_path = f"{path}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{extension}"
    workbook = xlrd.open_workbook(file_path)
    sheet = workbook.sheets()[0]
    rows_count = sheet.nrows
    title = sheet.row_values(0)[0]
    sheet_data = [sheet.row_values(row_index) for row_index in range(rows_count)]
    for title_start, operate in operator_matches.items():
        if title.strip().startswith(title_start):
            operate(sheet_data, template_path, output_path)


def write_project_excel(export_datas):
    print(export_datas)


if __name__ == "__main__":
    #     sample_file_path = 'basic/sample_export.xls'
    #     result = read_export_excel(file_path=sample_file_path)
    #     print(result)
    # file_path = sys.argv[1]
    # template_path = sys.argv[2]
    file_path, template_path = "basic/费用.xls", "basic/result_template.xlsx"
    result = read_export_excel(file_path, template_path)
    print(file_path, result)
